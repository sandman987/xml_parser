import xml.etree.ElementTree as ET

from columnar import columnar
from openpyxl import Workbook
import sys


def as_text(value):
    if value is None:
        return ""
    return str(value)

from fpdf import FPDF

path_to_file = input("file_path: ")

tree = ET.parse(path_to_file)

root = tree.getroot()

id_list = []
status_list = []
time_list = []
type_list = []
stream_list = []
description_list = []


lines_to_be_displayed = []

for i in range(len(root)):
    lines_to_be_displayed.append([])


#print(lines_to_be_displayed)
headers = ['#', 'Status', 'Time', 'Type', 'Stream', 'Description']


for alarm in root.findall('alarmLine'):
    id = alarm.attrib['id']
    status = alarm.attrib['status']
    time = alarm.attrib['time']
    type = alarm.attrib['type']
    stream = alarm.attrib['stream']
    description = alarm.attrib['description']

    id_list.append(id)
    status_list.append(status)
    time_list.append(time)
    type_list.append(type)
    stream_list.append(stream)
    description_list.append(description)


for i in range(len(lines_to_be_displayed)):
    lines_to_be_displayed[i] = [id_list[i], status_list[i], time_list[i], type_list[i], stream_list[i], description_list[i]]


table = columnar(lines_to_be_displayed, headers)
#print(table)

wb = Workbook()


for column in range(1, 7):
    wb.active.cell(row=1, column=column).value = headers[column - 1]

for i in range(len(lines_to_be_displayed)):
    for j in range(len(lines_to_be_displayed[i])):
        wb.active.cell(row=i + 2, column=j + 1).value = lines_to_be_displayed[i][j]

for column_cells in wb.active.columns:
    length = max(len(as_text(cell.value)) for cell in column_cells)
    wb.active.column_dimensions[column_cells[0].column_letter].width = length


wb.save("alarm_logs.xlsx")